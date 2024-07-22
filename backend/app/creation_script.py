import os
import openpyxl
import django
from django.conf import settings
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404

# Now you can import your models
from models import Department, CustomUser, Group

User = get_user_model()

# Load the workbook
workbook = openpyxl.load_workbook("accounts.xlsx")

# Assuming you want to work with the active sheet
worksheet = workbook.active

users = []

for row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    FN = row[0].value
    LN = row[1].value
    IN = row[2].value
    Dept = row[3].value
    Email = row[4].value
    Groups = row[5].value
    username = (FN + LN).lower()

    print(f"Creating user: {username}")
    print(f"First Name: {FN}")
    print(f"Last Name: {LN}")
    print(f"Initials: {IN}")
    print(f"Department: {Dept}")
    print(f"Email: {Email}")
    print(f"Groups: {Groups}")
    
    # Fetch Department instance or create if it doesn't exist
    department_instance, _ = Department.objects.get_or_create(name=Dept)

    # Create CustomUser instance
    custom_user = CustomUser(
        username=username,
        email=Email,
        first_name=FN,
        last_name=LN,
        initials=IN,
    )

    # Set groups
    if Groups:
        group_names = Groups.split(',')  # Assuming Groups are comma-separated
        groups = Group.objects.filter(name__in=group_names)
        print(f"Fetched groups: {groups}")
        custom_user.groups.set(list(groups))  # Convert to list before setting

    # Set department
    custom_user.department.add(department_instance)
    print(f"Created custom_user: {custom_user}")

    # Append the created CustomUser object to the list
    users.append(custom_user)

# Bulk create all users
CustomUser.objects.bulk_create(users)
